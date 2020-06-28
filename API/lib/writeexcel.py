# _*_ coding:utf-8 _*_
import os,sys
sys.path.append(os.path.dirname(os.path.dirname(__file__)))
import shutil
from API.config import setting
from openpyxl import load_workbook
from openpyxl.styles import Font,Alignment
from openpyxl.styles.colors import RED,GREEN,DARKYELLOW
import configparser as cparser
# --------- 读取config.ini配置文件 ---------------
cf = cparser.ConfigParser()
cf.read(setting.Test_Config,encoding='UTF-8')
name = cf.get("tester","name")

class WriteExcel():
    #-------------文件写入数据---------------
    def __init__(self,filename):
        self.filename = filename
        if not os.path.exists(self.filename):
            shutil.copyfile(setting.Source_File,setting.Target_File)#将文件1覆盖到文件2
        self.wb = load_workbook(self.filename)#打开文件
        self.ws = self.wb.active#获取第一个sheet

    def write_data(self,row_num,value):
        #写入测试结果
        font_Creen = Font(name='宋体',color= GREEN,bold=True)
        font_Red = Font(name='宋体',color=RED,bold=True)
        font_Yellow = Font(name='宋体',color=DARKYELLOW,bold=True)
        align = Alignment(horizontal='center',vertical='center')
        #获取所在行数
        L_num = "L"+str(row_num)
        M_num = "M"+str(row_num)
        print(L_num,M_num)
        if value =="PASS":
            self.ws.cell(row_num,12,value)#在第row_n行第12列写入value
            self.ws[L_num].font = font_Creen
        if value =="FAIL":
            self.ws.cell(row_num,12,value)
            self.ws[L_num].font = font_Red
        self.ws.cell(row_num,13,name)#在第row_n行第13列写入测试人员名字
        self.ws[L_num].alignment =align
        self.ws[M_num].alignment = align
        self.ws[M_num].font = font_Yellow
        self.wb.save(self.filename)
